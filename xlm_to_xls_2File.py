# CMD:      pip install  ne dela
# pycharm---- file----settings----interpreter----KNJIŽNICA---install package-------DONE
import xml.etree.ElementTree as ET
import xlwt
import os
#import pandas as pd
#import openpyxl
#from openpyxl import Workbook
import xlsxwriter as xlsxw
import win32com.client as win32
import pathlib

from pathlib import Path
from tkinter import filedialog
from tkinter import *
from xlwt import Workbook
import re
import time

now= time.strftime("%Y%m%d-%H%M%S")
output_filename = 'average_measurements_raw'+ now +'.xlsx'

workbook =xlsxw.Workbook(output_filename, {'strings_to_numbers': True})
sheet1 = workbook.add_worksheet()
bold = workbook.add_format({'bold': True})
cell_format_green = workbook.add_format()
cell_format_green.set_bg_color('green')
cell_format_red = workbook.add_format()
cell_format_red.set_bg_color('red')

sheet2 = workbook.add_worksheet()

root = Tk()
root.withdraw()
path = filedialog.askdirectory()

sheet1.write(0, 0, "Measurement path:",bold)
sheet1.write(0, 1, "DUT_Nr:",bold)
sheet1.write(0, 2, "Start time:",bold)
sheet1.write(0, 3, "Cfg path:",bold)
sheet1.write(0, 4, "Test result:",bold)
sheet1.write(0, 5, "Actual Speed[rpm]:",bold)
sheet1.write(0, 6, "Delta pressure[bar]:",bold)
sheet1.write(0, 7, "Efficiency[%]:",bold)
sheet1.write(0, 8, "Liquid flow[l/min]:",bold)
sheet1.write(0, 9, "Mechanical Power[W]:",bold)
sheet1.write(0, 10, "Power DUT[W]:",bold)
sheet1.write(0, 11, "Pressure in[bar]:",bold)
sheet1.write(0, 12, "Pressure out[bar]:",bold)
sheet1.write(0, 13, "Sense DUT[V]:",bold)
sheet1.write(0, 14, "Speed Request[rpm]:",bold)
sheet1.write(0, 15, "Supply current[A]:",bold)
sheet1.write(0, 16, "Supply power[W]:",bold)
sheet1.write(0, 17, "Supply voltage[V]:",bold)
sheet1.write(0, 18, "Temperature in[°C]:",bold)
sheet1.write(0, 19, "Temperature out[°C]:",bold)
sheet1.write(0, 20, "USM LEM current[A]:",bold)
sheet1.write(0, 21, "USM sense[V]:",bold)
sheet1.write(0, 22, "Valve value[°]:",bold)
sheet1.write(0, 23, "Power limit[W]:",bold)
sheet1.write(0, 24, "Order number:",bold)
sheet1.write(0, 25, "Comment:",bold)

i=1
j=0

for root, dirs, files in os.walk(path):
    for file in files:
        if file.endswith(".xml"):
             print(os.path.join(root, file))
             fullname = os.path.join(root, file)
             tree = ET.parse(fullname)
             xml = tree.getroot()
             sheet1.write_url(i, j, root)                                 #Measurement path
             j += 1
             for info in xml.findall('sut'):
                 dut = info.find("info[name='DUT']/description").text
                 sheet1.write(i, j, dut)                                  #DUT_Nr
                 print(dut)
                 j+=1
             for time in xml.findall('preparation'):
                 starttime = time.get('starttime')
                 sheet1.write(i, j, starttime)                            #Start time
                 print(starttime)
                 j += 1
             for xinfo in xml.findall('testsetup'):
                 cfg= xinfo.find("xinfo[name='Configuration']/description").text
                 sheet1.write(i, j, cfg)                                  #Cfg path
                 print(cfg)
                 j += 1
             for result in xml.findall('verdict'):
                 result = result.get('result')
                 if result == "pass":
                     sheet1.write(i, j, result,cell_format_green)
                 elif result == "fail":
                     sheet1.write(i, j, result,cell_format_red)
                 print(result)
             for measurements in xml.iter():
                 name = measurements.get('name')
                 value = measurements.get('value')
                 if name == "Speed Request:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Actual Speed:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Delta pressure:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Efficiency:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Liquid flow:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Mechanical Power:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Power DUT:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Pressure in:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Pressure out:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Sense DUT:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply current:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply power:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply voltage:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Temperature in:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Temperature out:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "USM LEM current:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "USM sense:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Valve value:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                     print('\n')
             for testcase in xml.findall('testgroup'):
                 if testcase.find("testcase[title='Input power limit']/teststep[1]") is not None:
                    powerlimit = testcase.find("testcase[title='Input power limit']/teststep[1]")
                    if powerlimit.text.find("measured input DC power") == -1:
                       powerlimit = testcase.find("testcase[title='Input power limit']/teststep[3]")
                    if powerlimit is not None:
                       powerlimit = powerlimit.text
                       j += 1
                       print(powerlimit)
                       value = powerlimit.split("measured input DC power", 1)[1]
                       value = value.split(" ", 2)[1]
                       sheet1.write(i, j, value)  # Power limit
                       break
             for info in xml.findall('sut'):
                 ordernr = info.find("info[name='Order number']/description")
                 if ordernr is not None:
                     ordernr = ordernr.text
                 else:
                     ordernr = None
                 j += 1
                 sheet1.write(i, j, ordernr)   #Order number
                 print(ordernr)
             for info in xml.findall('sut'):
                 comment = info.find("info[name='Comment']/description")
                 if comment is not None:
                     comment = comment.text
                 else:
                     comment = None
                 j += 1
                 sheet1.write(i, j, comment)                              #Comment
                 print(comment)
             j=0
             i+=1


chart1= workbook.add_chart({'type':'column'})
chart1.add_series({'values' : ['Sheet1', 1, 5, i-1, 5],
                   'name'  : 'Actual Speed'
                   })

chart1.set_title({'name':'Actual Speed'})
sheet2.insert_chart('A1',chart1)

chart2= workbook.add_chart({'type':'column'})
chart2.add_series({'values': ['Sheet1', 1, 6, i-1, 6],
                   'name'  : 'Delta pressure'
                   })

chart2.set_title({'name':'Delta pressure'})
sheet2.insert_chart('N1',chart2)

chart3= workbook.add_chart({'type':'column'})
chart3.add_series({'values': ['Sheet1', 1, 8, i-1, 8],
                   'name'  : 'Liquid flow'
                   })

chart3.set_title({'name':'Liquid flow'})
sheet2.insert_chart('A23',chart3)

chart4= workbook.add_chart({'type':'column'})
chart4.add_series({'values': ['Sheet1', 1, 10, i-1, 10],
                   'name'  : 'Power DUT'
                   })

chart4.set_title({'name':'Power DUT'})
sheet2.insert_chart('N23',chart4)

workbook.close()


excel = win32.gencache.EnsureDispatch('Excel.Application')
dirname = pathlib.Path().absolute()
suffix = '.xlsx'
filename = output_filename
final_path = Path(dirname, filename).with_suffix(suffix)
print("Directory Path:", final_path)
wb = excel.Workbooks.Open(final_path)
ws = wb.Worksheets("Sheet1")
ws.Columns.AutoFit()
wb.Save()
excel.Application.Quit()








































































output_filename = 'average_measurements'+ now + '.xlsx'

workbook =xlsxw.Workbook(output_filename, {'strings_to_numbers': True})
sheet1 = workbook.add_worksheet()
bold = workbook.add_format({'bold': True})
cell_format_green = workbook.add_format()
cell_format_green.set_bg_color('green')
cell_format_red = workbook.add_format()
cell_format_red.set_bg_color('red')

sheet2 = workbook.add_worksheet()

root = Tk()
root.withdraw()
#path = filedialog.askdirectory()

sheet1.write(0, 0, "Measurement path:",bold)
sheet1.write(0, 1, "DUT_Nr:",bold)
sheet1.write(0, 2, "Start time:",bold)
sheet1.write(0, 3, "Cfg path:",bold)
sheet1.write(0, 4, "Test result:",bold)
sheet1.write(0, 5, "Actual Speed[rpm]:",bold)
sheet1.write(0, 6, "Delta pressure[bar]:",bold)
sheet1.write(0, 7, "Efficiency[%]:",bold)
sheet1.write(0, 8, "Liquid flow[l/min]:",bold)
sheet1.write(0, 9, "Mechanical Power[W]:",bold)
sheet1.write(0, 10, "Power DUT[W]:",bold)
sheet1.write(0, 11, "Pressure in[bar]:",bold)
sheet1.write(0, 12, "Pressure out[bar]:",bold)
sheet1.write(0, 13, "Sense DUT[V]:",bold)
sheet1.write(0, 14, "Speed Request[rpm]:",bold)
sheet1.write(0, 15, "Supply current[A]:",bold)
sheet1.write(0, 16, "Supply power[W]:",bold)
sheet1.write(0, 17, "Supply voltage[V]:",bold)
sheet1.write(0, 18, "Temperature in[°C]:",bold)
sheet1.write(0, 19, "Temperature out[°C]:",bold)
sheet1.write(0, 20, "USM LEM current[A]:",bold)
sheet1.write(0, 21, "USM sense[V]:",bold)
sheet1.write(0, 22, "Valve value[°]:",bold)
sheet1.write(0, 23, "Power limit[W]:",bold)
sheet1.write(0, 24, "Order number:",bold)
sheet1.write(0, 25, "Comment:",bold)

i=1
j=0

for root, dirs, files in os.walk(path):
    for file in files:
        if file.endswith(".xml"):
             print(os.path.join(root, file))
             fullname = os.path.join(root, file)
             tree = ET.parse(fullname)
             xml = tree.getroot()
             sheet1.write_url(i, j, root)                                 #Measurement path
             j += 1
             for info in xml.findall('sut'):
                 dut = info.find("info[name='DUT']/description").text
                 sheet1.write(i, j, dut)                                  #DUT_Nr
                 print(dut)
                 j+=1
             for time in xml.findall('preparation'):
                 starttime = time.get('starttime')
                 sheet1.write(i, j, starttime)                            #Start time
                 print(starttime)
                 j += 1
             for xinfo in xml.findall('testsetup'):
                 cfg= xinfo.find("xinfo[name='Configuration']/description").text
                 sheet1.write(i, j, cfg)                                  #Cfg path
                 print(cfg)
                 j += 1
             for result in xml.findall('verdict'):
                 result = result.get('result')
                 if result == "pass":
                     sheet1.write(i, j, result,cell_format_green)
                 elif result == "fail":
                     sheet1.write(i, j, result,cell_format_red)
                 print(result)
             for measurements in xml.iter():
                 name = measurements.get('name')
                 value = measurements.get('value')
                 if name == "Speed Request:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Actual Speed:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Delta pressure:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Efficiency:":
                     j += 1
                     sheet1.write(i,j,value)
                     print(name, value)
                 elif name == "Liquid flow:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Mechanical Power:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Power DUT:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Pressure in:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Pressure out:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Sense DUT:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply current:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply power:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Supply voltage:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Temperature in:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Temperature out:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "USM LEM current:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "USM sense:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                 elif name == "Valve value:":
                     j += 1
                     sheet1.write(i, j, value)
                     print(name, value)
                     print('\n')
             for testcase in xml.findall('testgroup'):
                 if testcase.find("testcase[title='Input power limit']/teststep[1]") is not None:
                    powerlimit = testcase.find("testcase[title='Input power limit']/teststep[1]")
                    if powerlimit.text.find("measured input DC power") == -1:
                       powerlimit = testcase.find("testcase[title='Input power limit']/teststep[3]")
                    if powerlimit is not None:
                       powerlimit = powerlimit.text
                       j += 1
                       print(powerlimit)
                       value = powerlimit.split("measured input DC power", 1)[1]
                       value = value.split(" ", 2)[1]
                       sheet1.write(i, j, value)  # Power limit
                       break
             for info in xml.findall('sut'):
                 ordernr = info.find("info[name='Order number']/description")
                 if ordernr is not None:
                     ordernr = ordernr.text
                 else:
                     ordernr = None
                 j += 1
                 sheet1.write(i, j, ordernr)   #Order number
                 print(ordernr)
             for info in xml.findall('sut'):
                 comment = info.find("info[name='Comment']/description")
                 if comment is not None:
                     comment = comment.text
                 else:
                     comment = None
                 j += 1
                 sheet1.write(i, j, comment)                              #Comment
                 print(comment)
             j=0
             i+=1


chart1= workbook.add_chart({'type':'column'})
chart1.add_series({'values' : ['Sheet1', 1, 5, i-1, 5],
                   'name'  : 'Actual Speed'
                   })

chart1.set_title({'name':'Actual Speed'})
sheet2.insert_chart('A1',chart1)

chart2= workbook.add_chart({'type':'column'})
chart2.add_series({'values': ['Sheet1', 1, 6, i-1, 6],
                   'name'  : 'Delta pressure'
                   })

chart2.set_title({'name':'Delta pressure'})
sheet2.insert_chart('N1',chart2)

chart3= workbook.add_chart({'type':'column'})
chart3.add_series({'values': ['Sheet1', 1, 8, i-1, 8],
                   'name'  : 'Liquid flow'
                   })

chart3.set_title({'name':'Liquid flow'})
sheet2.insert_chart('A23',chart3)

chart4= workbook.add_chart({'type':'column'})
chart4.add_series({'values': ['Sheet1', 1, 10, i-1, 10],
                   'name'  : 'Power DUT'
                   })

chart4.set_title({'name':'Power DUT'})
sheet2.insert_chart('N23',chart4)

workbook.close()







excel = win32.gencache.EnsureDispatch('Excel.Application')
dirname = pathlib.Path().absolute()
suffix = '.xlsx'
filename = output_filename
final_path = Path(dirname, filename).with_suffix(suffix)
print("Directory Path:", final_path)
wb = excel.Workbooks.Open(final_path)
ws = wb.Worksheets("Sheet1")
ws.Columns.AutoFit()
wb.Save()
excel.Application.Quit()


filename = output_filename
sheetname = 'Sheet1'
xl = win32.DispatchEx('Excel.Application')

path =  os.getcwd().replace('\'','\\') + '\\'
wb = xl.Workbooks.Open(path + filename)

#wb = xl.Workbooks.Open(Filename=filename)
ws = wb.Sheets(sheetname)

def filter():
    without_bad_lines = True
    while without_bad_lines == True:
        without_bad_lines = False
        nrows = ws.UsedRange.Rows.Count
        for i in range(2, nrows+1):

            # define emptiness of cell
            speed = ws.Cells(i, 6).Value
            result = ws.Cells(i, 5).Value
            power_limit = ws.Cells(i, 24).Value
            try:
                bad_values = int(speed) >= 10000 or int(speed) <= 100 or int(power_limit) >= 1000 or int(
                    power_limit) <= 200
                if result == 'fail' or bad_values:
                    # Delete indexes of rows
                    ws.Cells(i, 1).EntireRow.Delete()
                    without_bad_lines = True

            except:
                ws.Cells(i, 1).EntireRow.Delete()
                without_bad_lines = True
    wb.Save()


filter()





item_list = []
without_bad_lines = True
while without_bad_lines == True:
    without_bad_lines = False
    nrows = ws.UsedRange.Rows.Count
    for i in range(2, nrows):
        DUT_N = ws.Cells(i, 2).Value
        seen = set(item_list)

        if DUT_N not in seen:
            seen.add(DUT_N)
            item_list.append(DUT_N)
        else:
            ws.Cells(i, 1).EntireRow.Delete()
            without_bad_lines = True
#vzame tisto z najboljsim efficiency, torej odstrani ostale
wb.Save()



nrows = ws.UsedRange.Rows.Count
ncolumns = ws.UsedRange.Columns.Count
from openpyxl import load_workbook
book = load_workbook(filename = output_filename)

shit = book.get_active_sheet()
shit.cell(row=nrows+2, column=5).value = "average"
shit.cell(row=nrows+3, column=5).value = "-∆(AVG-MIN) "
shit.cell(row=nrows+4, column=5).value = "+∆ (MAX-AVG)"
shit.cell(row=nrows+5, column=5).value = "-% "      
shit.cell(row=nrows+6, column=5).value = "+% "



all = 0
max_value = -10000000.0
min_value = +10000000.0
for col in range(6, ncolumns-1):
    for i in range(2, nrows+1):
        value = ws.Cells(i, col).Value
        #print (value)
        all = all + float(value)
        if max_value <= value:
            max_value = value
        if min_value >= value:
            min_value = value
    average = float(all)/float(nrows-1)
    shit.cell(row=i+2, column=col).value = average
    shit.cell(row=i + 3, column=col).value = average - min_value
    shit.cell(row=i + 4, column=col).value = max_value - average
    shit.cell(row=i + 5, column=col).value = (average - min_value)/average
    shit.cell(row=i + 6, column=col).value = (max_value - average)/ average

    max_value = -10000000.0
    min_value = +10000000.0
    all  = 0
wb.Save()
wb.Close()
book.save(output_filename)
xl.Quit()
"""



     
    

  
"""